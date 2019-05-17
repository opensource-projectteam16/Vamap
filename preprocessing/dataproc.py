import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
import csv
from io import StringIO
import numpy.lib.recfunctions as npfunc
from pathlib import Path
import platform
# Seokcheon Ju is in charge

''' 
<How to work>

<Usage> 
 Input example  :
 Output example :

'''

class Coordinate:
    def __init__(self,xcor,ycor):
        self.x=xcor
        self.y=ycor

class dataproc:
    base_route=""
    user_route=""
    tab=""
    coverage=0
    user_list=[]
    base_data=[]
    def __init__(self,datalist,mode):
        """ 
        fileroute={[list of datapath],a datapath}, columnname=['x_columnname','y_columnname','x"_columnname','y"_columnname','value_columnname'], sheet=sheet name(if different)
        if it is road or double coordinate data road=true
        mode means recall from ready made data. 0==no csv file 1==have csv file
        """
        system=platform.system()

        if system=='Linux':
            self.base_route="/base_data"
            self.user_route="/user_data"
            self.tab="/"
        elif system=='Windows':
            self.base_route="\\base_data"
            self.user_route="\\user_data"
            self.tab="\\"
            
        if mode==0:
            self.base_data=datalist
        if mode==1:
            self.base_data=datalist
        if mode==2:
            for data in datalist:
                self.base_data.append(data)

        mypath = os.path.realpath(__file__)
        fullPath=os.path.dirname(os.path.dirname(mypath))
        finalpath=""
        for parent in os.listdir(fullPath):
            #print(parent+"  ")
            if parent[-4:] == 'data':
                finalpath=os.path.join(fullPath,parent)
        if mode==0:
            fullPath=finalpath+""+self.user_route
        else:
            fullPath=finalpath+""+self.base_route
        print("fulli"+fullPath)
        if os.path.isdir(fullPath):
            print ("initialize")
            filelist=[]
            for difile in os.listdir(fullPath):
                for datafile in self.base_data:
                    if mode==2 and datafile[0]==difile:
                        filelist.append(fullPath+self.tab+difile)
                    elif datafile==difile:
                        filelist.append(fullPath+self.tab+difile)
         
            datamanager=excelmanager(filelist,self.base_data,mode)
            self.maindata=datamanager.getdata()
            self.datalabel=datamanager.getdatalabel()
            self.changedtype(mode)

    def changedtype(self,mode):
        if mode==0:
            print('hi')
            for index in range(0,len(self.maindata)+1):
                print('hi1')
                if index==0:
                    self.maindata[0][index].dtype.names='x'
                elif index==1:
                    self.maindata[0][index].dtype.names='y'
                elif index==2:
                    self.maindata[0][index].dtype.names='value'
                else:
                    self.maindata[0][index].dtype.names='string'
        if mode==1:
            for index in range(0,len(self.maindata)+1):
                if index==0:
                    self.maindata[0][index].dtype.names='x'
                elif index==1:
                    self.maindata[0][index].dtype.names='y'
                elif index==2:
                    self.maindata[0][index].dtype.names='x1'
                elif index==3:
                    self.maindata[0][index].dtype.names='y1'
                elif index==4:
                    self.maindata[0][index].dtype.names='x2'
                elif index==5:
                    self.maindata[0][index].dtype.names='y2'
                elif index==6:
                    self.maindata[0][index].dtype.names='value'
                else:
                    self.maindata[0][index].dtype.names='string'
        if mode==2:
            for datalist in self.maindata:
                for index in zip(datalist,range(0,len(datalist)+1)):
                    if index==0:
                        datalist[0][index].dtype.names='x'
                    elif index==1:
                        datalist[0][index].dtype.names='y'
                    elif index==2:
                        datalist[0][index].dtype.names='value'
                    else:
                        datalist[0][index].dtype.names='string'
                
    def changecolumnname(self,columnname):
        if len(columnname):
            pass
    
    def savedata(self):
        writer = csv.writer(open("datalabel.csv", 'w'))
        for row in self.datalabel:
            writer.writerow(row)
        for datpage, datlabel in zip(self.maindata,self.datalabel):
            filename=datlabel+'.csv'
            np.save(filename,datpage)           
    def getdatalabel(self):
        return self.datalabel
    def getdata(self):
        return self.maindata

class excelmanager:
    def __init__(self,datafile,basedata,mode): 
        self.resultlist=[]
        self.datalabel=[]
        print(datafile,basedata)
        if mode==2:
            for afile,datacolumn in zip(datafile,basedata):
                datalist=datacolumn[2:]
                print("loadstart"+afile,datalist)
                load_exs = load_workbook(filename=afile, data_only=True)
                print("loaded "+str(afile))
                for sheet in load_exs: #several sheets
                    if sheet.title==datacolumn[1]:
                        print(sheet,datalist)
                        self.resultlist.append(self.extractdata(sheet,datalist))
                        self.datalabel.append(sheet.title)

        else:
            for afile in datafile:
                datalist=basedata[2:]
                print("loadstart"+afile,datalist)
                load_exs = load_workbook(filename=afile, data_only=True)
                print("loaded "+str(afile))
                for sheet in load_exs: #several sheets
                    if sheet.title==basedata[1]:
                        print('163',sheet.title,datalist)
                        self.resultlist.append(self.extractdata(sheet,datalist))
                        self.datalabel.append(sheet.title)            
    
    def extractdata(self,load_sheet,columnname):
        result=[]
        namelist=[]
        typelist=[]
        listtype=0
        for p,r in zip(load_sheet[2],load_sheet[1]):
            for name in columnname:
                if r.value==name:
                    if isinstance(p.value,str):
                        typelist.append(object)
                        namelist.append(name)
                    else:
                        typelist.append(np.float64)
                        namelist.append(name)

        dat=[]
        for name in namelist:
            for r in load_sheet[1]:
                if r.value==name:
                    dtype=0
                    tempdat=np.array([row[r.column-1].value for row in load_sheet.iter_rows(min_row=2)],order='K')[np.newaxis]
                    tempdat=tempdat.astype([(name,typelist[listtype])])
                    tempdat=tempdat.T
                    dat.append(tempdat)
                    listtype=listtype+1
        return dat

    def getdata(self):
        return self.resultlist

    def getdatalabel(self):
        return self.datalabel
