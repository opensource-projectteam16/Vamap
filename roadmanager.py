# import pandas as pd ## IMPORT ERROR


import sys
import os
import numpy as np

from geopy.geocoders import Nominatim
from openpyxl import load_workbook


# Hyunjae Lee is in charge

'''
<How to work>

<Usage>
 Input example  :
        python roadmanager.py <your_roadfile.xlsx> <sheet_name> <column_name>
 Output example :
        It will return (latitude, longitude) in the address
'''


def main():

    checkInputs(sys.argv)

    # Get user inputs
    your_roadfile = sys.argv[1]
    sheet_name = sys.argv[2]
    column_name = sys.argv[3]

    # Change the address to (latitude, longitude)
    your_roadfile, sheet_name = returnColumn(
        your_roadfile, sheet_name, column_name, returntype=1)

    print(your_roadfile, sheet_name)


def changeAddress(geolocator, address):

    location = geolocator.geocode(address)

    if location is None:
        print("No (latitude, longitude) is available at the address {}".format(
            address))
        return 0, 0

    return location.latitude, location.longitude


def checkInputs(checkInput):
    if len(checkInput) != 4:
        print(checkInput)
        printError()


def returnColumn(path_your_roadfile, sheet_name,
                 column_name, startrow=1, returntype=0):
    '''
    returntype = 0 : return fp, sheet_name, column_name
    returntype = 1 : return fp, sheet_name
    '''
    # If returntype is 1
    if returntype == 1:
        geolocator = Nominatim(user_agent=" THIS IS ROAD MANAGER", timeout=10)

    # <your_roadfile.csv> <sheet_name> <column_name>

    ''' 1st check '''
    # if your_roadfile is vaild form
    if os.path.isfile(path_your_roadfile):
        print("Correct access")
    else:
        print("[ERROR] Failed to load Excel File : %s" % (path_your_roadfile))

    ''' 2nd check '''
    # if your_roadfile has the given 'sheet_name'
    # Execl handler
    wb = load_workbook(filename=path_your_roadfile)
    if wb is None:
        printError()

    # get the worksheet
    worksheet = wb[sheet_name]
    print(worksheet)
    if worksheet is None:
        printError()

    ''' 3rd check '''
    # if the worksheet has the given 'column_name'
    # Check columns
    list_with_values = []

    for cell in worksheet[startrow]:
        list_with_values.append(cell.value)

    print('Your column list is {}'.format(list_with_values))
    if column_name not in list_with_values:
        printError()

    if returntype == 0:
        wb.close()
        return path_your_roadfile, sheet_name, column_name

    ''' 4th check '''
    # Get the contents of the given column

    column_num = list_with_values.index(column_name) + 1
    changed_latitude_list = list()
    changed_longitude_list = list()

    print('{} is located at {} column'.format(column_name, column_num))
    print('worksheet.max_row :', worksheet.max_row)
    # Get the value of the column and change it to coordinate value

    for r in range(2, worksheet.max_row+1):

        d = worksheet.cell(row=r, column=column_num)
        changed_latitude, changed_longitude = changeAddress(
            geolocator, d.value)

        if d.value is None:
            print('There is no value at {} row , {} column'.format(r, column_num))
            break

        changed_latitude_list.append(changed_latitude)
        changed_longitude_list.append(changed_longitude)
        print(r)
        print(d.value)
        print(changed_latitude, changed_longitude)

    ''' 5th check '''
    # Write each list to a new column
    initialized_row = 2
    initialized_col = len(list_with_values)+1

    print('Changing Addresses is about to start !')
    for each_latitude in changed_latitude_list:
        worksheet.cell(row=initialized_row,
                       column=initialized_col).value = each_latitude
        initialized_row += 1

    worksheet.cell(row=1, column=initialized_col).value = column_name + '_x'
    initialized_col += 1
    initialized_row = 2

    for each_longitude in changed_longitude_list:
        worksheet.cell(row=initialized_row,
                       column=initialized_col).value = each_longitude
        initialized_row += 1

    worksheet.cell(row=1, column=initialized_col).value = column_name + '_y'

    print('Changing Addresses is finished ! ')

    # Return and save the file
    if returntype == 1:
        # save data as excel file
        wb.save(filename='converted_address.xlsx')
        wb.close()
        print(" Finish to save !")
        return path_your_roadfile, sheet_name
    else:
        printError()


def printError():
    print("You have to follow input rule")
    print("python roadmanager.py <your_roadfile.csv> <sheet_name> <column_name>")
    quit()


if __name__ == "__main__":
    main()
