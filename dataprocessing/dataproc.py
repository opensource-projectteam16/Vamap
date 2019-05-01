import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
import csv
from io import StringIO

# Seokcheon Ju is in charge

''' 
<How to work>

<Usage> 
 Input example  :
 Output example :

'''

class Coordinate:
    def __init__(self,xcor,ycor):
        self.x=xcor
        self.y=ycor

class dataproc: 
    def __init__(self,fileroute,columnx='x',columny='y',column_x='_x',column_y='_y',value='value',road=False,sheetname='',mode=0,):
        """ 
        fileroute={[list of datapath],a datapath}, columnx,column_x=name of x axis, columny,column_y=name of y axis, 
        value= column which will be major comparator, sheet=sheet name(if different)
        if it is road or double coordinate data road=true
        mode means recall from ready made data. 0==no csv file 1==have csv file
        PLEASE ADD ROAD DATA SEPERATELY
        """
        if mode==0:
            datamanager=excelmanager(fileroute,columnx,columny,column_x,column_y,value,sheetname,road)
            self.maindata=datamanager.getdata()
            self.datalabel=datamanager.getdatalabel()
           # print (self.maindata)
           # print (self.datalabel)
            self.savedata()
        if mode==1:
            self.datalabel=[]
            with open("datalabel.csv", 'r') as csvFile:
                reader = csv.reader(csvFile)
                for row in reader:
                    if row != []:
                        self.datalabel.append(''.join(row))
                csvFile.close()
            self.maindata=[]
            for difile in os.listdir(fileroute):
               # print(difile)
                if difile.split(".")[-1]=='csv' and difile!='datalabel.csv':
                    data = np.genfromtxt(difile, delimiter=",")
                    #print(data)
                    self.maindata.append(data)
                    csvFile.close()
        

    def getcoordsortedata(self,coord1,coord2,coord3,coord4):
        """
        starts from right top coordinate, rotate into CW
        """
        resultlist=[]
        for datpage in self.maindata:
            idx = np.zeros((1,))
            for row in range(0,np.size(datpage,0)):
                tmp=np.where(np.logical_and(datpage[row][0] > coord2.y , datpage[row][0] <= coord1.y))[0]
                if tmp.size>0:
                    idx=np.concatenate((idx,[1]))
                else:
                    idx=np.concatenate((idx,[0]))
                
            idx=np.delete(idx,0,0)        
            xindex=list()
            for r in range(0,np.size(idx,0)):
                if idx[r]!=0:
                    xindex.append(r)
            #extracted x < x> coordinate        
            idx = np.zeros((1,))
            for row in xindex:
                tmp=np.where(np.logical_and(datpage[row][1] >coord3.x,datpage[row][1] <=coord2.x))[0]
                if tmp.size>0:
                    idx=np.concatenate((idx,[1]))
                else:
                    idx=np.concatenate((idx,[0]))
            idx=np.delete(idx,0,0)            
            xyindex=list()
            for r in range(0,np.size(idx,0)):
                if idx[r]!=0:
                    xyindex.append(xindex[r])    
            resultlist.append(datpage[np.array(np.asarray(xyindex))])      
        
        return resultlist
    
    def savedata(self):
        writer = csv.writer(open("datalabel.csv", 'w'))
        for row in self.datalabel:
            writer.writerow(row)
        for datpage, datlabel in zip(self.maindata,self.datalabel):
            filename=datlabel+'.csv'
            np.savetxt(filename, datpage, delimiter=',')

    def getdatalabel(self):
        return self.datalabel
    def getdata(self):
        return self.maindata

class excelmanager:
    def __init__(self,datafile,columnx,columny,column_x,column_y,value,sheetname='',road=False): 
        if road==False:
            columnname=[columnx,columny,value]
            self.resultlist=[]
            self.datalabel=[]

            if type(datafile)==list:
                for afile in datafile:
                    load_exs = load_workbook(filename=afile, data_only=True)
                    print("loaded "+str(afile))
                    if sheetname!='':
                        load_ws = load_exs[sheetname]
                        self.datalabel.append(sheetname)
                        self.resultlist=extractdata(load_exs[sheetname],columnname,road)

                    for sheet in load_exs: #several sheets
                        self.resultlist.append(self.extractdata(sheet,columnname,road))
                        self.datalabel.append(sheet.title)

            else:
                load_exl = load_workbook(filename=datafile, data_only=True)
                print("loaded "+str(datafile))

                if sheetname!='':
                    load_ws = load_exl[sheetname]
                    self.datalabel.append(sheetname)
                    self.resultlist=extractdata(load_ws[sheetname],columnname,road)

                for sheet in load_exl: #several sheets
                    self.resultlist.append(self.extractdata(sheet,columnname,road))
                    self.datalabel.append(sheet.title)

        if road==True:
            columnname=[columnx,columny,column_x,column_y,value]
            self.resultlist=[]
            self.datalabel=['road']
            load_exl = load_workbook(filename=datafile, data_only=True)
            print("loaded "+str(datafile))
            if sheetname!='':
                load_ws = load_exl[sheetname]
                self.resultlist=extractdata(load_ws[sheetname],columnname,road)

            for sheet in load_exl: #several sheets
                self.resultlist.append(self.extractdata(sheet,columnname,road))
        
    
    def extractdata(self,load_sheet,columnname,road):
        if road==False:
            result=np.zeros((load_sheet.max_row-1,1))
            for name in columnname:
                for r in load_sheet[1]:
                    if r.value==name:
                        dat=np.array([row[r.column-1].value for row in load_sheet.iter_rows(min_row=2)],order='K')
                        dat=dat[:,np.newaxis]
                result=np.concatenate((result,dat),1)
            result=np.delete(result,0,1)
        else:
            result=np.zeros((load_sheet.max_row-1,1))
            for name in columnname:
                for r in load_sheet[1]:
                    if r.value==name:
                        dat=np.array([row[r.column-1].value for row in load_sheet.iter_rows(min_row=2)],order='K')
                        dat=dat[:,np.newaxis]
                result=np.concatenate((result,dat),1)
            result=np.delete(result,0,1)
        return result

    def getdata(self):
        return self.resultlist
    def getdatalabel(self):
        return self.datalabel
